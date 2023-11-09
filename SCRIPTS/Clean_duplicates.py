import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Spécifiez le chemin complet vers le fichier Excel
input_file = './BDD_steps/BDD_Car_MAJ.xlsx'

output_file = "./BDD_steps/Clean_duplicates.xlsx"
non_integer_file_c4cID = "./BDD_steps/NonIntegerC4Cid.csv"
non_integer_file_tdID = "./BDD_steps/NonIntegerTdID.csv"
debug_file = "./BDD_steps/Clean_duplicates.csv"

# Charger le fichier Excel en spécifiant la feuille
df = pd.read_excel(input_file, sheet_name='BDD_2')

# Affiche la liste des feuilles dans le fichier d'entrée
print(f"Feuilles dans le fichier d'entrée : {', '.join(pd.ExcelFile(input_file).sheet_names)}")

# Gestion des doublons dans la colonne CAR_C4C_ID
c4cID_duplicates_before = df[df.duplicated(subset='car_c4cID', keep='first')]
if not c4cID_duplicates_before.empty:
    c4cID_duplicates_before.to_csv('car_c4cID_duplicates_before.csv', index=False)
else:
    print("Aucun doublon trouvé dans la colonne 'car_c4cID' avant les modifications.")

df['car_c4cID'] = pd.to_numeric(df['car_c4cID'], errors='coerce')

empty_values = df['car_c4cID'].isna()
duplicates = df[df.duplicated(subset='car_c4cID', keep='first')]

max_id = df.loc[~empty_values, 'car_c4cID'].max()
if pd.notna(max_id):
    max_id = int(max_id)
else:
    max_id = 0

for i, row in df.iterrows():
    if empty_values[i] or not empty_values[i] and i in duplicates.index:
        max_id += 1
        df.at[i, 'car_c4cID'] = max_id

# Comptage du nombre de lignes avant le traitement de CAR_TDID
num_rows_before = df.shape[0]

# Enregistrez le fichier modifié dans le fichier de sortie
df.to_excel(output_file, sheet_name='BDD_2', index=False)

# Comptez le nombre de lignes dans le fichier de sortie
num_output_rows = len(df)

# Ouvrez le fichier Excel en utilisant openpyxl
book = load_workbook(input_file)

# Sélectionnez la feuille 'BDD_2'
ws = book['BDD_2']

# Supprimez les données existantes dans la colonne 'car_c4cID'
for row in ws.iter_rows(min_row=2, min_col=2):
    for cell in row:
        cell.value = None

# Écrivez les données du dataframe dans la feuille 'BDD_2' sans écraser les autres feuilles
for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
    for col_index, value in enumerate(row, start=1):
        ws.cell(row=row_index, column=col_index, value=value)

# Enregistrez le fichier mis à jour
book.save(input_file)

# Affiche la liste des feuilles après la mise à jour
updated_sheets = book.sheetnames
print(f"Feuilles après la mise à jour : {', '.join(updated_sheets)}")

print(f"La colonne 'car_c4cID' de la feuille 'BDD_2' a été mise à jour dans le fichier '{input_file}'.")
print(f"Le fichier '{input_file}' a été converti en '{output_file}'")
print(f"Nombre de lignes dans le fichier de sortie : {num_output_rows}")